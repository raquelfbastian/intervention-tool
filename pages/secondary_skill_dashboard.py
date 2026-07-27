import io
import pandas as pd
import streamlit as st
import plotly.express as px
from io import BytesIO
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# CONFIG
# ============================================================

DEFAULT_SKILLS_FILE = "input/Dump_20_7_2026.xlsx"
DEFAULT_TARGET_FILE = "input/career_level_targets.xlsx"
DEFAULT_PROJECT_FILE = "input/HC Report ao 072126 1.xlsx"

ALLOWED_BUSINESS_GROUPS = [
    "Tech_Song",
    "Tech_Adobe Platform",
]

SKILL_TYPE_FILTER = "Secondary"
DEFAULT_MIN_PROJECT_RESOURCES = 5

# Skills dump columns
COL_RESOURCE_ID = "Personnel No"
COL_ENTERPRISE_ID = "Enterpriseid"
COL_CAREER_LEVEL_FROM_DUMP = "Career level"
COL_SKILL_NAME = "SkillName"
COL_BUSINESS_GROUP = "Business Group"
COL_PROFICIENCY = "proficiency"
COL_SKILL_TYPE = "Skill type"

# Project lookup columns
PROJECT_PEOPLEKEY_COL = "Personnel No"
PROJECT_NAME_COL = "Project"


# Target reference columns
TARGET_LEVEL_CODE_COL = "Career Level Code"
TARGET_LEVEL_COL = "Career Level"
TARGET_PROFICIENCY_COL = "Target Proficiency"

# Project lookup columns
PROJECT_PEOPLEKEY_COL = "Personnel No"
PROJECT_NAME_COL = "Project"
PROJECT_BUSINESS_GROUP_COL = "Business Group"

REQUIRED_SKILLS_COLUMNS = [
    COL_RESOURCE_ID,
    COL_ENTERPRISE_ID,
    COL_CAREER_LEVEL_FROM_DUMP,
    COL_SKILL_NAME,
    COL_BUSINESS_GROUP,
    COL_PROFICIENCY,
    COL_SKILL_TYPE,
]

REQUIRED_TARGET_COLUMNS = [
    TARGET_LEVEL_CODE_COL,
    TARGET_LEVEL_COL,
    TARGET_PROFICIENCY_COL,
]

REQUIRED_PROJECT_COLUMNS = [
    PROJECT_PEOPLEKEY_COL,
    PROJECT_NAME_COL,
]


# ============================================================
# HELPERS
# ============================================================

def normalize_col_name(value):
    return str(value).strip().lower().replace(" ", "")


def clean_column_names(df):
    df.columns = [str(c).strip() for c in df.columns]
    return df


def source_to_excel_io(source):
    """
    Accepts either a local file path or uploaded file bytes.
    Returns something pandas can read.
    """
    if isinstance(source, bytes):
        return io.BytesIO(source)
    return source


def read_excel_normal(source):
    return pd.read_excel(source_to_excel_io(source))


def find_column(df, expected_name):
    expected = normalize_col_name(expected_name)

    for col in df.columns:
        if normalize_col_name(col) == expected:
            return col

    return None


def validate_columns(df, required_columns, file_label):
    missing = []

    for col in required_columns:
        actual_col = find_column(df, col)
        if actual_col is None:
            missing.append(col)

    if missing:
        raise ValueError(
            f"{file_label} is missing required columns: {missing}. "
            f"Available columns: {list(df.columns)}"
        )


def read_excel_detect_header(source, required_headers, file_label):
    """
    Reads an Excel file where headers may not be on row 1.
    Detects the row containing all required headers.
    Works with both local paths and uploaded file bytes.
    """
    raw = pd.read_excel(source_to_excel_io(source), header=None)

    required_normalized = [normalize_col_name(h) for h in required_headers]

    header_row_index = None

    for idx, row in raw.iterrows():
        row_values = [normalize_col_name(x) for x in row.tolist()]

        if all(req in row_values for req in required_normalized):
            header_row_index = idx
            break

    if header_row_index is None:
        raise ValueError(
            f"Could not detect header row for {file_label}. "
            f"Required headers: {required_headers}"
        )

    df = pd.read_excel(source_to_excel_io(source), header=header_row_index)
    df = clean_column_names(df)

    validate_columns(df, required_headers, file_label)

    return df


def normalize_peoplekey(value):
    if pd.isna(value):
        return None

    try:
        return str(int(float(value))).strip()
    except Exception:
        return str(value).strip()


def parse_number(value):
    if pd.isna(value):
        return None

    try:
        return int(float(value))
    except Exception:
        return None


def parse_proficiency(value):
    if pd.isna(value):
        return None

    text = str(value).strip().upper()

    if text in ["", "NULL", "NAN", "NONE"]:
        return None

    if text.startswith("P"):
        text = text.replace("P", "").strip()

    try:
        return int(float(text))
    except Exception:
        return None


def target_label(value):
    if pd.isna(value):
        return "No Target"

    return f"P{int(value)}"


def proficiency_label(value):
    if pd.isna(value):
        return "No Assessment"

    try:
        value = int(value)

        proficiency_map = {
            -1: "P0",
            0: "P1",
            1: "P2",
            2: "P3",
            3: "Expert Eligible",
        }

        return proficiency_map.get(value, f"P{value}")

    except Exception:
        return "No Assessment"


def safe_pct(numerator, denominator):
    if denominator == 0:
        return 0

    return numerator / denominator * 100


def format_pct(value):
    return f"{value:.1f}%"


def action_reason(row):
    if not row["has_assessment"]:
        return "No Assessment"

    if row["below_target"]:
        return "Below Target"

    return "Meeting Target"


def get_source(uploaded_file, default_path):
    """
    Uses uploaded file when provided; otherwise falls back to local default.
    """
    if uploaded_file is None:
        return default_path

    # If it's a Streamlit UploadedFile or similar buffer-like object
    if hasattr(uploaded_file, "getvalue"):
        return uploaded_file.getvalue()

    # If it's already a bytes or a file path string, return as-is
    return uploaded_file


def distinct_count_where(df, condition, id_col):
    return df.loc[condition, id_col].nunique()



def create_chase_excel(
    summary_df,
    project_df,
    resource_detail_df,
    skill_gap_df,
    career_summary_df,
    selected_business_group,
    assessment_scope,
):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ------------------------------------------------------------
        # Write sheets
        # ------------------------------------------------------------
        summary_df.to_excel(
            writer,
            sheet_name="Executive Summary",
            index=False,
            startrow=0,
        )

        project_df.to_excel(
            writer,
            sheet_name="Project Action Summary",
            index=False,
        )

        resource_detail_df.to_excel(
            writer,
            sheet_name="Resource Chase Detail",
            index=False,
        )

        skill_gap_df.to_excel(
            writer,
            sheet_name="Skill Gap Summary",
            index=False,
        )

        career_summary_df.to_excel(
            writer,
            sheet_name="Career Level Health",
            index=False,
        )

        workbook = writer.book
        ws = writer.sheets["Executive Summary"]
        from openpyxl.styles import Font, PatternFill, Alignment

        # Style Executive Summary row 1 and row 2
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        value_fill = PatternFill(
            fill_type="solid",
            fgColor="F3F8FC"
        )

        for cell in ws[1]:
            cell.font = Font(
                bold=True,
                size=13
            )
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        for cell in ws[2]:
            cell.font = Font(
                bold=True,
                size=13
            )
            cell.fill = value_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 24


        # ------------------------------------------------------------
        # Add metric definitions in same Executive Summary sheet
        # ------------------------------------------------------------
        definitions = [
            ["Metric", "Meaning"],
            ["Business Group", "Business group included in the analysis, such as Tech_Song, Tech_Adobe Platform, or All."],
            ["Assessment Scope", "Indicates whether the analysis includes Primary Skills, Secondary Skills, or All Skills."],
            ["Total Resources", "Total unique resources included in the selected business group and assessment scope."],
            ["Assessed Resources", "Resources with a completed competency assessment."],
            ["Completion %", "Percentage of resources with completed assessments. Formula: Assessed Resources / Total Resources x 100."],
            ["Target Compliance %", "Percentage of resources meeting or exceeding the required target proficiency."],
            ["No Assessment", "Resources without a completed competency assessment. These require assessment completion follow-up."],
            ["Below Target", "Resources who completed an assessment but are below the required target proficiency. These require capability uplift, learning, coaching, or reassessment action."],
        ]

        # Put definitions starting row 5 para hindi matatamaan yung summary
        start_row = 5

        for r_idx, row in enumerate(definitions, start=start_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # ------------------------------------------------------------
        # Create Top Projects chart as image
        # ------------------------------------------------------------
        chart_start_row = start_row + len(definitions) + 3

        try:
            top_projects = (
                project_df
                .sort_values("Resources To Chase", ascending=False)
                .head(10)
                .copy()
            )

            if len(top_projects) > 0:
                plt.figure(figsize=(10, 6))
                plt.barh(
                    top_projects["Project"],
                    top_projects["Resources To Chase"],
                )
                plt.xlabel("Resources To Chase")
                plt.ylabel("Project")
                plt.title("Top Projects With Most Resources To Chase")
                plt.gca().invert_yaxis()
                plt.tight_layout()

                chart_path = "top_projects_chart.png"
                plt.savefig(chart_path, dpi=150)
                plt.close()

                img = ExcelImage(chart_path)
                img.width = 720
                img.height = 420

                ws.add_image(img, f"A{chart_start_row}")

        except Exception:
            # Do not fail Excel generation if chart creation fails
            ws.cell(
                row=chart_start_row,
                column=1,
                value="Chart could not be generated. Please refer to Project Action Summary tab.",
            )

        # ------------------------------------------------------------
        # Basic formatting
        # ------------------------------------------------------------
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            if sheet_name == "Resource Chase Detail":
                project_header_fill = PatternFill(
                    fill_type="solid",
                    fgColor="BDD7EE"
                )

                for row in range(2, worksheet.max_row + 1):
                    project_value = worksheet.cell(row=row, column=2).value

                    if isinstance(project_value, str) and project_value.startswith("PROJECT:"):
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.font = Font(bold=True, size=12)
                            cell.fill = project_header_fill
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                        worksheet.row_dimensions[row].height = 22
            worksheet.freeze_panes = "A2"

            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    try:
                        cell_length = len(str(cell.value)) if cell.value is not None else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except Exception:
                        pass

                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)

    output.seek(0)
    # Ensure Enterprise IDs in exported Excel are lowercase where present
    try:
        if "Enterpriseid" in resource_detail_df.columns:
            resource_detail_df["Enterpriseid"] = resource_detail_df["Enterpriseid"].astype(str).str.lower()
    except Exception:
        pass

    return output

def build_grouped_resource_export(resource_export):
    grouped_rows = []
    columns = list(resource_export.columns)

    for project_name, group in resource_export.groupby("Project", sort=True):
        project_header = {col: "" for col in columns}
        project_header["Project"] = f"PROJECT: {project_name}"
        grouped_rows.append(project_header)

        grouped_rows.extend(group.to_dict("records"))

    return pd.DataFrame(grouped_rows, columns=columns)


# ============================================================
# DATA PIPELINE
# ============================================================

@st.cache_data(show_spinner=False)
def build_data(skills_source, target_source, project_source, selected_business_group):

    allowed_bg_upper = [bg.upper() for bg in ALLOWED_BUSINESS_GROUPS]

    # ---------------------------
    # Load skills dump
    # ---------------------------
    skills_df = read_excel_normal(skills_source)
    skills_df = clean_column_names(skills_df)
    validate_columns(skills_df, REQUIRED_SKILLS_COLUMNS, "Skills Dump")

    actual_resource_col = find_column(skills_df, COL_RESOURCE_ID)
    actual_career_col = find_column(skills_df, COL_CAREER_LEVEL_FROM_DUMP)
    actual_skill_col = find_column(skills_df, COL_SKILL_NAME)
    actual_bg_col = find_column(skills_df, COL_BUSINESS_GROUP)
    actual_prof_col = find_column(skills_df, COL_PROFICIENCY)
    actual_skill_type_col = find_column(skills_df, COL_SKILL_TYPE)
    actual_enterprise_id_col = find_column(skills_df, COL_ENTERPRISE_ID)

    skills_df = skills_df.rename(
        columns={
            actual_resource_col: COL_RESOURCE_ID,
            actual_enterprise_id_col: COL_ENTERPRISE_ID,
            actual_career_col: COL_CAREER_LEVEL_FROM_DUMP,
            actual_skill_col: COL_SKILL_NAME,
            actual_bg_col: COL_BUSINESS_GROUP,
            actual_prof_col: COL_PROFICIENCY,
            actual_skill_type_col: COL_SKILL_TYPE,
        }
    )

    # Normalize Enterprise IDs to lowercase for uniform display across the app
    skills_df[COL_ENTERPRISE_ID] = (
        skills_df[COL_ENTERPRISE_ID]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.lower()
    )

    skills_df["Business Group Clean"] = (
        skills_df[COL_BUSINESS_GROUP]
        .astype(str)
        .str.strip()
    )

    if selected_business_group == "All":
        skills_df = skills_df[
            skills_df["Business Group Clean"]
            .str.upper()
            .isin(allowed_bg_upper)
        ].copy()
    else:
        skills_df = skills_df[
            skills_df["Business Group Clean"]
            .str.upper()
            == selected_business_group.upper()
        ].copy()

    if len(skills_df) == 0:
        raise ValueError(
            f"Skills Dump has no records after Business Group filter: "
            f"{selected_business_group}"
        )

    skills_df["Skill Type Clean"] = (
        skills_df[COL_SKILL_TYPE]
        .astype(str)
        .str.strip()
    )

    # Business rule for this page: selected business group(s) first, then Secondary skill type only.
    skills_df = skills_df[
        skills_df["Skill Type Clean"].str.upper() == SKILL_TYPE_FILTER.upper()
    ].copy()

    if len(skills_df) == 0:
        raise ValueError(
            f"No rows found after Skill Type filter: {SKILL_TYPE_FILTER}"
        )

    # ---------------------------
    # Load target lookup
    # ---------------------------
    target_df = read_excel_detect_header(
        target_source,
        REQUIRED_TARGET_COLUMNS,
        "Career Level Target Lookup"
    )

    actual_target_code_col = find_column(target_df, TARGET_LEVEL_CODE_COL)
    actual_target_level_col = find_column(target_df, TARGET_LEVEL_COL)
    actual_target_prof_col = find_column(target_df, TARGET_PROFICIENCY_COL)

    target_df["career_level_code"] = target_df[actual_target_code_col].apply(parse_number)
    target_df["career_level_num"] = target_df[actual_target_level_col].apply(parse_number)
    target_df["target_proficiency_num"] = target_df[actual_target_prof_col].apply(parse_proficiency)

    target_lookup = target_df.loc[
        target_df["career_level_code"].notna()
        & target_df["career_level_num"].notna()
        & target_df["target_proficiency_num"].notna(),
        [
            "career_level_code",
            "career_level_num",
            "target_proficiency_num",
        ],
    ].drop_duplicates(subset=["career_level_code"]).copy()

    # ---------------------------
    # Clean skills data and map targets
    # ---------------------------
    skills_df[COL_RESOURCE_ID] = skills_df[COL_RESOURCE_ID].apply(normalize_peoplekey)
    skills_df["career_level_code"] = skills_df[COL_CAREER_LEVEL_FROM_DUMP].apply(parse_number)
    skills_df["proficiency_num"] = skills_df[COL_PROFICIENCY].apply(parse_proficiency)

    skills_df = skills_df.merge(
        target_lookup,
        on="career_level_code",
        how="left",
    )

    skills_df["has_assessment"] = skills_df["proficiency_num"].notna()

    skills_df["meets_target"] = (
        skills_df["proficiency_num"].notna()
        & skills_df["target_proficiency_num"].notna()
        & (skills_df["proficiency_num"] >= skills_df["target_proficiency_num"])
    )

    # IMPORTANT:
    # Below target is assessed resources only.
    # No assessment is handled separately.
    skills_df["below_target"] = (
        skills_df["proficiency_num"].notna()
        & skills_df["target_proficiency_num"].notna()
        & (skills_df["proficiency_num"] < skills_df["target_proficiency_num"])
    )

    # ---------------------------
    # Build resource-level view
    # ---------------------------
    resource_df = (
        skills_df
        .groupby(COL_RESOURCE_ID, as_index=False)
        .agg(
            business_group=("Business Group Clean", "first"),
            career_level_code=("career_level_code", "first"),
            career_level_num=("career_level_num", "first"),
            primary_skill=(COL_SKILL_NAME, "first"),
            max_proficiency_num=("proficiency_num", "max"),
            target_proficiency_num=("target_proficiency_num", "first"),
            has_assessment=("has_assessment", "max"),
            Enterpriseid=("Enterpriseid", "first")
        )
    )

    resource_df["meets_target"] = (
        resource_df["max_proficiency_num"].notna()
        & resource_df["target_proficiency_num"].notna()
        & (resource_df["max_proficiency_num"] >= resource_df["target_proficiency_num"])
    )

    resource_df["below_target"] = (
        resource_df["max_proficiency_num"].notna()
        & resource_df["target_proficiency_num"].notna()
        & (resource_df["max_proficiency_num"] < resource_df["target_proficiency_num"])
    )

    resource_df["Target"] = resource_df["target_proficiency_num"].apply(target_label)
    resource_df["Actual"] = resource_df["max_proficiency_num"].apply(proficiency_label)
    resource_df["Action Reason"] = resource_df.apply(action_reason, axis=1)

    # ---------------------------
    # Load project lookup
    # ---------------------------
    project_df = read_excel_detect_header(
        project_source,
        REQUIRED_PROJECT_COLUMNS,
        "Project Lookup"
    )

    actual_project_peoplekey_col = find_column(project_df, PROJECT_PEOPLEKEY_COL)
    actual_project_name_col = find_column(project_df, PROJECT_NAME_COL)

    project_df = project_df.rename(
        columns={
            actual_project_peoplekey_col: PROJECT_PEOPLEKEY_COL,
            actual_project_name_col: PROJECT_NAME_COL,
        }
    )

    project_df[PROJECT_PEOPLEKEY_COL] = project_df[PROJECT_PEOPLEKEY_COL].apply(normalize_peoplekey)

    project_lookup = (
        project_df[
            [
                PROJECT_PEOPLEKEY_COL,
                PROJECT_NAME_COL,
            ]
        ]
        .dropna(subset=[PROJECT_PEOPLEKEY_COL])
        .drop_duplicates(subset=[PROJECT_PEOPLEKEY_COL])
        .copy()
    )

    resource_project_df = resource_df.merge(
        project_lookup,
        left_on=COL_RESOURCE_ID,
        right_on=PROJECT_PEOPLEKEY_COL,
        how="left",
    )

    resource_project_df["Project"] = resource_project_df[PROJECT_NAME_COL].fillna("Unmapped")

    # ---------------------------
    # Project-level view
    # ---------------------------
    project_summary_rows = []

    for project_name, group in resource_project_df.groupby("Project"):
        total_resources = group[COL_RESOURCE_ID].nunique()
        assessed_resources = distinct_count_where(
            group,
            group["has_assessment"] == True,
            COL_RESOURCE_ID
        )
        no_assessment = distinct_count_where(
            group,
            group["has_assessment"] == False,
            COL_RESOURCE_ID
        )
        below_target_only = distinct_count_where(
            group,
            (group["below_target"] == True) & (group["has_assessment"] == True),
            COL_RESOURCE_ID
        )
        meeting_target = distinct_count_where(
            group,
            group["meets_target"] == True,
            COL_RESOURCE_ID
        )

        resources_to_chase = no_assessment + below_target_only

        project_summary_rows.append(
            {
                "Project": project_name,
                "TotalResources": total_resources,
                "AssessedResources": assessed_resources,
                "No Assessment": no_assessment,
                "Below Target Only": below_target_only,
                "Meeting Target": meeting_target,
                "Resources To Chase": resources_to_chase,
                "Chase %": safe_pct(resources_to_chase, total_resources),
                "Completion %": safe_pct(assessed_resources, total_resources),
                "Target Compliance %": safe_pct(meeting_target, total_resources),
                "Priority Score": no_assessment + (below_target_only * 2),
            }
        )

    project_view = pd.DataFrame(project_summary_rows)

    if len(project_view) > 0:
        project_view = project_view.sort_values(
            ["Resources To Chase", "TotalResources"],
            ascending=[False, False]
        )

    metadata = {
        "skills_rows_after_filters": len(skills_df),
        "target_mappings": len(target_lookup),
        "project_lookup_peoplekeys": project_lookup[PROJECT_PEOPLEKEY_COL].nunique(),
        "unmapped_resources": resource_project_df[
            resource_project_df["Project"] == "Unmapped"
        ][COL_RESOURCE_ID].nunique(),
    }

    return skills_df, resource_df, resource_project_df, project_view, metadata


# ============================================================


def render_secondary_dashboard(
    skills_source=None,
    target_source=None,
    project_source=None,
    selected_business_group="All",
    show_business_group_select=True,
    set_page_config=True,
):
    if set_page_config:
        st.set_page_config(
            page_title="myCompetency Secondary Skills Explorer",
            page_icon="📌",
            layout="wide",
        )

    #st.title("🚀 myCompetency Secondary Skills Explorer")
    #st.caption("Secondary Skill capability and future-readiness analytics.")

    st.sidebar.header("Data Upload")
    st.sidebar.caption("Upload the 3 Excel files. If blank, app uses local files under input.")

    # If the caller provided uploaded file objects (Streamlit UploadedFile),
    # the renderer will use them; otherwise show upload widgets here.
    if skills_source is None:
        uploaded_skills = st.sidebar.file_uploader(
            "1. Skills Dump",
            type=["xlsx"],
            key="skills_upload",
        )
    else:
        uploaded_skills = skills_source

    if target_source is None:
        uploaded_target = st.sidebar.file_uploader(
            "2. Career Level Target Lookup",
            type=["xlsx"],
            key="target_upload",
        )
    else:
        uploaded_target = target_source

    if project_source is None:
        uploaded_project = st.sidebar.file_uploader(
            "3. Project Lookup",
            type=["xlsx"],
            key="project_upload",
        )
    else:
        uploaded_project = project_source

    # Convert uploaded objects to sources usable by the data pipeline
    skills_source = get_source(uploaded_skills, DEFAULT_SKILLS_FILE)
    target_source = get_source(uploaded_target, DEFAULT_TARGET_FILE)
    project_source = get_source(uploaded_project, DEFAULT_PROJECT_FILE)

    if show_business_group_select:
        business_group_options = ["All"] + ALLOWED_BUSINESS_GROUPS

        selected_business_group = st.sidebar.selectbox(
            "Business Group",
            business_group_options,
            index=0,
        )

    assessment_scope = SKILL_TYPE_FILTER

    scorecard_scope = (
        "Tech_Song + Tech_Adobe Platform"
        if selected_business_group == "All"
        else selected_business_group
    )

    with st.sidebar.expander("Expected Columns", expanded=False):
        st.markdown("**Skills Dump**")
        st.code("\n".join(REQUIRED_SKILLS_COLUMNS))

        st.markdown("**Career Level Target Lookup**")
        st.code("\n".join(REQUIRED_TARGET_COLUMNS))

        st.markdown("**Project Lookup**")
        st.code("\n".join(REQUIRED_PROJECT_COLUMNS))

    # ============================================================
    # BUILD DATA
    # ============================================================

    try:
        skills_df, resource_df, resource_project_df, project_view, metadata = build_data(
            skills_source,
            target_source,
            project_source,
            selected_business_group,
        )

        st.sidebar.success("Data loaded and validated")
        st.sidebar.caption(f"Business Group: {scorecard_scope}")
        st.sidebar.caption(f"Filtered skill rows: {metadata['skills_rows_after_filters']:,}")
        st.sidebar.caption(f"Target mappings: {metadata['target_mappings']:,}")
        st.sidebar.caption(f"Project lookup peoplekeys: {metadata['project_lookup_peoplekeys']:,}")
        st.sidebar.caption(f"Unmapped resources: {metadata['unmapped_resources']:,}")

    except Exception as e:
        st.error("Failed to load or validate input files.")
        st.exception(e)
        st.stop()

    # ============================================================
    # DASHBOARD FILTERS
    # ============================================================

    st.sidebar.header("Dashboard Filters")

    project_options = ["All"] + sorted(
        resource_project_df["Project"]
        .dropna()
        .unique()
        .tolist()
    )
    selected_project = st.sidebar.selectbox("Project", project_options)

    career_options = ["All"] + sorted(
        [
            int(x)
            for x in resource_project_df["career_level_num"]
            .dropna()
            .unique()
            .tolist()
        ],
        reverse=True,
    )
    selected_career = st.sidebar.selectbox("Career Level", career_options)

    skill_options = ["All"] + sorted(
        resource_project_df["primary_skill"]
        .dropna()
        .unique()
        .tolist()
    )
    selected_skill = st.sidebar.selectbox("Secondary Skill", skill_options)

    secondary_detail = resource_project_df.copy()

    if selected_project != "All":
        secondary_detail = secondary_detail[
            secondary_detail["Project"] == selected_project
        ]

    if selected_career != "All":
        secondary_detail = secondary_detail[
            secondary_detail["career_level_num"] == selected_career
        ]

    if selected_skill != "All":
        secondary_detail = secondary_detail[
            secondary_detail["primary_skill"] == selected_skill
        ]

    # ============================================================
    # PRIMARY VS SECONDARY CAPABILITY ALIGNMENT
    # ============================================================

    st.subheader("Primary vs Secondary Capability Alignment")

    st.caption(
        "This view compares target attainment across Primary and Secondary skills. "
        "Use this for capability planning only, not individual performance evaluation."
    )

    # ------------------------------------------------------------
    # Reload raw skills dump so we have BOTH Primary and Secondary
    # ------------------------------------------------------------

    all_skills_df = read_excel_normal(skills_source)
    all_skills_df = clean_column_names(all_skills_df)

    actual_resource_col = find_column(all_skills_df, COL_RESOURCE_ID)
    actual_career_col = find_column(all_skills_df, COL_CAREER_LEVEL_FROM_DUMP)
    actual_skill_col = find_column(all_skills_df, COL_SKILL_NAME)
    actual_bg_col = find_column(all_skills_df, COL_BUSINESS_GROUP)
    actual_prof_col = find_column(all_skills_df, COL_PROFICIENCY)
    actual_skill_type_col = find_column(all_skills_df, COL_SKILL_TYPE)

    all_skills_df = all_skills_df.rename(
        columns={
            actual_resource_col: COL_RESOURCE_ID,
            actual_career_col: COL_CAREER_LEVEL_FROM_DUMP,
            actual_skill_col: COL_SKILL_NAME,
            actual_bg_col: COL_BUSINESS_GROUP,
            actual_prof_col: COL_PROFICIENCY,
            actual_skill_type_col: COL_SKILL_TYPE,
        }
    )

    all_skills_df[COL_RESOURCE_ID] = all_skills_df[COL_RESOURCE_ID].apply(normalize_peoplekey)

    all_skills_df["Business Group Clean"] = (
        all_skills_df[COL_BUSINESS_GROUP]
        .astype(str)
        .str.strip()
    )

    all_skills_df["Skill Type Clean"] = (
        all_skills_df[COL_SKILL_TYPE]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    allowed_bg_upper = [bg.upper() for bg in ALLOWED_BUSINESS_GROUPS]

    if selected_business_group == "All":
        all_skills_df = all_skills_df[
            all_skills_df["Business Group Clean"]
            .str.upper()
            .isin(allowed_bg_upper)
        ].copy()
    else:
        all_skills_df = all_skills_df[
            all_skills_df["Business Group Clean"]
            .str.upper()
            == selected_business_group.upper()
        ].copy()

    all_skills_df = all_skills_df[
        all_skills_df["Skill Type Clean"].isin(["PRIMARY", "SECONDARY"])
    ].copy()

    target_df = read_excel_detect_header(
        target_source,
        REQUIRED_TARGET_COLUMNS,
        "Career Level Target Lookup"
    )

    actual_target_code_col = find_column(target_df, TARGET_LEVEL_CODE_COL)
    actual_target_level_col = find_column(target_df, TARGET_LEVEL_COL)
    actual_target_prof_col = find_column(target_df, TARGET_PROFICIENCY_COL)

    target_df["career_level_code"] = target_df[actual_target_code_col].apply(parse_number)
    target_df["career_level_num"] = target_df[actual_target_level_col].apply(parse_number)
    target_df["target_proficiency_num"] = target_df[actual_target_prof_col].apply(parse_proficiency)

    target_lookup = target_df.loc[
        target_df["career_level_code"].notna()
        & target_df["career_level_num"].notna()
        & target_df["target_proficiency_num"].notna(),
        [
            "career_level_code",
            "career_level_num",
            "target_proficiency_num",
        ],
    ].drop_duplicates(subset=["career_level_code"]).copy()

    all_skills_df["career_level_code"] = all_skills_df[COL_CAREER_LEVEL_FROM_DUMP].apply(parse_number)
    all_skills_df["proficiency_num"] = all_skills_df[COL_PROFICIENCY].apply(parse_proficiency)

    all_skills_df = all_skills_df.merge(
        target_lookup,
        on="career_level_code",
        how="left",
    )

    all_skills_df["has_assessment"] = all_skills_df["proficiency_num"].notna()

    all_skills_df["meets_target"] = (
        all_skills_df["proficiency_num"].notna()
        & all_skills_df["target_proficiency_num"].notna()
        & (all_skills_df["proficiency_num"] >= all_skills_df["target_proficiency_num"])
    )

    all_skills_df["below_target"] = (
        all_skills_df["proficiency_num"].notna()
        & all_skills_df["target_proficiency_num"].notna()
        & (all_skills_df["proficiency_num"] < all_skills_df["target_proficiency_num"])
    )

    primary_secondary_status = (
    all_skills_df
    .groupby(COL_RESOURCE_ID)
    .apply(
        lambda g: pd.Series(
            {
                "Has Primary": (
                    g["Skill Type Clean"]
                    .eq("PRIMARY")
                    .any()
                ),
                "Has Secondary": (
                    g["Skill Type Clean"]
                    .eq("SECONDARY")
                    .any()
                ),

                "Primary Meets Target": (
                    g.loc[
                        g["Skill Type Clean"].eq("PRIMARY"),
                        "meets_target"
                    ]
                    .any()
                ),
                "Primary Below Target": (
                    g.loc[
                        g["Skill Type Clean"].eq("PRIMARY"),
                        "below_target"
                    ]
                    .any()
                ),

                "Secondary Meets Target": (
                    g.loc[
                        g["Skill Type Clean"].eq("SECONDARY"),
                        "meets_target"
                    ]
                    .any()
                ),
                "Secondary Below Target": (
                    g.loc[
                        g["Skill Type Clean"].eq("SECONDARY"),
                        "below_target"
                    ]
                    .any()
                ),

                # Include Enterpriseid (EID) so we can display it in tables
                "EID": (
                    g[COL_ENTERPRISE_ID]
                    .dropna()
                    .astype(str)
                    .unique()
                ) and (g[COL_ENTERPRISE_ID].dropna().astype(str).unique()[0]) or "",

                "Primary Skills": ", ".join(
                    sorted(
                        g.loc[
                            g["Skill Type Clean"].eq("PRIMARY"),
                            COL_SKILL_NAME
                        ]
                        .dropna()
                        .astype(str)
                        .unique()
                    )
                ),

                "Secondary Skills Meeting Target": ", ".join(
                    sorted(
                        g.loc[
                            g["Skill Type Clean"].eq("SECONDARY")
                            & g["meets_target"].eq(True),
                            COL_SKILL_NAME
                        ]
                        .dropna()
                        .astype(str)
                        .unique()
                    )
                ),
            }
        )
    )
    .reset_index()
)

    def classify_primary_secondary(row):
        if row["Primary Meets Target"] and row["Secondary Meets Target"]:
            return "Meets Primary + Meets Secondary"

        if row["Primary Below Target"] and row["Secondary Meets Target"]:
            return "Below Primary Target + Meets Secondary Target"

        if row["Primary Meets Target"] and row["Secondary Below Target"]:
            return "Meets Primary + Below Secondary Target"

        if row["Primary Below Target"] and row["Secondary Below Target"]:
            return "Below Primary Target + Below Secondary Target"

        if row["Primary Meets Target"] and not row["Has Secondary"]:
            return "Meets Primary Only"

        if row["Primary Below Target"] and not row["Has Secondary"]:
            return "Below Primary Target Only"

        if not row["Has Primary"] and row["Secondary Meets Target"]:
            return "Meets Secondary Only"

        return "Other / Needs Review"

    primary_secondary_status["Capability Alignment"] = (
        primary_secondary_status.apply(
            classify_primary_secondary,
            axis=1
        )
    )

    primary_secondary_summary = (
        primary_secondary_status
        .groupby("Capability Alignment", as_index=False)
        .agg(
            Resources=(COL_RESOURCE_ID, "nunique")
        )
        .sort_values(
            "Resources",
            ascending=False
        )
    )

    st.dataframe(
        primary_secondary_summary,
        width="stretch",
        hide_index=True,
    )

    fig_alignment = px.bar(
        primary_secondary_summary.sort_values("Resources", ascending=True),
        x="Resources",
        y="Capability Alignment",
        orientation="h",
        title="Primary vs Secondary Capability Alignment",
    )

    fig_alignment.update_layout(
        height=450,
        yaxis_title="Capability Alignment",
        xaxis_title="Resources",
    )

    st.plotly_chart(
        fig_alignment,
        width="stretch",
    )

    below_primary_pass_secondary = primary_secondary_status[
        primary_secondary_status["Capability Alignment"]
        == "Below Primary Target + Meets Secondary Target"
    ].copy()

    st.subheader("Below Primary Target + Meets Secondary Target")

    st.caption(
        "Resources below target in at least one primary skill but meeting target in at least one secondary skill. "
        "This can indicate adjacent capability strengths for planning or redeployment discussions."
    )

    st.metric(
        "Resources",
        f"{below_primary_pass_secondary['EID'].nunique():,}"
    )

    st.dataframe(
        below_primary_pass_secondary[
            [
                "EID",
                "Primary Skills",
                "Secondary Skills Meeting Target",
                "Capability Alignment",
            ]
        ],
        width="stretch",
        hide_index=True,
    )

    secondary_skill_records = len(secondary_detail)

    secondary_unique_resources = (
        secondary_detail[COL_RESOURCE_ID]
        .dropna()
        .nunique()
    )

    secondary_unique_skills = (
        secondary_detail["primary_skill"]
        .dropna()
        .nunique()
    )

    secondary_proficient_resources = (
        secondary_detail.loc[
            secondary_detail["meets_target"] == True,
            COL_RESOURCE_ID,
        ]
        .nunique()
    )

    secondary_below_target_resources = (
        secondary_detail.loc[
            secondary_detail["below_target"] == True,
            COL_RESOURCE_ID,
        ]
        .nunique()
    )

    secondary_not_assessed_resources = max(
        secondary_unique_resources
        - secondary_proficient_resources
        - secondary_below_target_resources,
        0,
    )

    avg_secondary_skills_per_resource = (
        round(secondary_skill_records / secondary_unique_resources, 1)
        if secondary_unique_resources > 0
        else 0
    )

    st.subheader("Secondary Skill Overview")
    st.caption(
        "This page filters the skills dump by the selected Business Group and Skill Type = Secondary. "
        "Resource counts are distinct Personnel No values, while skill records count the secondary skill rows."
    )

    row1_col1, row1_col2, row1_col3 = st.columns(3)
    row1_col1.metric("Resources with Secondary Skills", f"{secondary_unique_resources:,}")
    row1_col2.metric("Unique Secondary Skills", f"{secondary_unique_skills:,}")
    row1_col3.metric("Avg Secondary Skills / Resource", avg_secondary_skills_per_resource)

    row2_col1, row2_col2, row2_col3 = st.columns(3)
    row2_col1.metric("Proficient Resources", f"{secondary_proficient_resources:,}")
    row2_col3.metric("Not Yet Assessed", f"{secondary_not_assessed_resources:,}")

    st.subheader("Top Secondary Skills")
    st.caption(
        "Shows secondary skill coverage and how many distinct resources are meeting, below, or missing the expected target proficiency."
    )

    if len(secondary_detail) == 0:
        st.warning("No secondary skill records found for the current filters.")
    else:
        secondary_skill_summary = (
            secondary_detail
            .groupby("primary_skill")
            .apply(
                lambda g: pd.Series(
                    {
                        "Resources": g[COL_RESOURCE_ID].nunique(),
                        "Resources Meeting Target": g.loc[
                            g["meets_target"] == True,
                            COL_RESOURCE_ID,
                        ].nunique(),
                        "Resources Below Target": g.loc[
                            g["below_target"] == True,
                            COL_RESOURCE_ID,
                        ].nunique(),
                        "Assessed Resources": g.loc[
                            g["has_assessment"] == True,
                            COL_RESOURCE_ID,
                        ].nunique(),
                    }
                )
            )
            .reset_index()
        )

        secondary_skill_summary["Resources Not Assessed"] = (
            secondary_skill_summary["Resources"]
            - secondary_skill_summary["Resources Meeting Target"]
            - secondary_skill_summary["Resources Below Target"]
        ).clip(lower=0)

        secondary_skill_summary["Target Compliance %"] = (
            secondary_skill_summary["Resources Meeting Target"]
            / secondary_skill_summary["Resources"]
            * 100
        ).round(0)

        secondary_skill_summary["Below Target %"] = (
            secondary_skill_summary["Resources Below Target"]
            / secondary_skill_summary["Resources"]
            * 100
        ).round(0)

        secondary_skill_summary = secondary_skill_summary.rename(
            columns={
                "primary_skill": "Secondary Skill",
            }
        )

        secondary_skill_summary = secondary_skill_summary.sort_values(
            "Resources",
            ascending=False,
        )

        top_secondary_skills = secondary_skill_summary.head(30)

        fig_secondary = px.bar(
            top_secondary_skills.sort_values("Resources", ascending=True),
            x="Resources",
            y="Secondary Skill",
            orientation="h",
            title="Top Secondary Skills by Resource Coverage",
            hover_data=[
                "Resources Meeting Target",
                "Target Compliance %",
                "Resources Below Target",
                "Below Target %",
                "Resources Not Assessed",
            ],
        )

        fig_secondary.update_layout(
            height=700,
            yaxis_title="Secondary Skill",
            xaxis_title="Resources",
        )

        st.plotly_chart(
            fig_secondary,
            width="stretch",
        )

        st.dataframe(
            top_secondary_skills[
                [
                    "Secondary Skill",
                    "Resources",
                    "Resources Meeting Target",
                    "Target Compliance %",
                    "Resources Below Target",
                    "Below Target %",
                    "Resources Not Assessed",
                ]
            ],
            width="stretch",
            hide_index=True,
        )

    return {
        "skills_df": skills_df,
        "resource_df": resource_df,
        "resource_project_df": resource_project_df,
        "project_view": project_view,
        "metadata": metadata,
        "secondary_detail": secondary_detail,
    }


if __name__ == "__main__":
    render_secondary_dashboard()
