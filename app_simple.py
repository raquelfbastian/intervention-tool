import io
import pandas as pd
import streamlit as st
import plotly.express as px
import io
import re

from secondary_skill_dashboard import render_secondary_dashboard

# ============================================================
# CONFIG
# ============================================================

DEFAULT_MASTER_FILE = "input/MyC_Report_as_of_2026_07_22_Tech_.xlsx"
DEFAULT_RESULT_FILE = "input/Dump_27_07_2026.xlsx"

DETAILS_SHEET_NAME = "Details"

ALLOWED_BUSINESS_GROUPS = [
    "Tech_Song",
    "Tech_Adobe Platform",
]

COL_PERSONNEL_NO = "Personnel No"
COL_EID = "Enterpriseid"
COL_LEVEL = "Management Level"
COL_SKILL_NAME = "SkillName"
COL_SKILL_TYPE = "Skill type"
COL_BUSINESS_GROUP = "Business Group"
COL_PROJECT = "Project Name"

RESULT_COL_EID = "Enterpriseid"
RESULT_COL_SKILL = "SkillName"
RESULT_COL_PROFICIENCY = "proficiency"
RESULT_COL_PROFICIENCY_DESC = "Proficiency Description"


# ============================================================
# HELPERS
# ============================================================

def source_to_excel_io(source):
    if isinstance(source, bytes):
        return io.BytesIO(source)
    return source


def clean_column_names(df):
    df.columns = [str(c).strip() for c in df.columns]
    return df


def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_key(value):
    if pd.isna(value):
        return ""

    text = str(value).strip()

    try:
        return str(int(float(text)))
    except Exception:
        return text


def read_master_file(source):
    df = pd.read_excel(
        source_to_excel_io(source),
        sheet_name=DETAILS_SHEET_NAME,
        header=1,
    )

    df = clean_column_names(df)

    required = [
        COL_PERSONNEL_NO,
        COL_EID,
        COL_LEVEL,
        COL_SKILL_NAME,
        COL_SKILL_TYPE,
        COL_BUSINESS_GROUP,
        COL_PROJECT,
    ]

    missing = [
        col for col in required
        if col not in df.columns
    ]

    if missing:
        raise ValueError(
            f"Master file missing columns: {missing}. "
            f"Available columns: {list(df.columns)}"
        )

    return df


def read_result_file(source):
    df = pd.read_excel(source_to_excel_io(source))
    df = clean_column_names(df)

    required = [
        RESULT_COL_EID,
        RESULT_COL_SKILL,
        RESULT_COL_PROFICIENCY,
    ]

    missing = [
        col for col in required
        if col not in df.columns
    ]

    if missing:
        raise ValueError(
            f"Result file missing columns: {missing}. "
            f"Available columns: {list(df.columns)}"
        )

    return df


@st.cache_data
def build_data(master_source, result_source, selected_business_group, selected_skill_type):
    # ---------------------------
    # Load master
    # ---------------------------
    master_df = read_master_file(master_source)

    master_df[COL_PERSONNEL_NO] = master_df[COL_PERSONNEL_NO].apply(normalize_key)
    master_df[COL_EID] = master_df[COL_EID].apply(normalize_text).str.lower()
    master_df[COL_SKILL_NAME] = master_df[COL_SKILL_NAME].apply(normalize_text)
    master_df[COL_SKILL_TYPE] = master_df[COL_SKILL_TYPE].apply(normalize_text)
    master_df[COL_BUSINESS_GROUP] = master_df[COL_BUSINESS_GROUP].apply(normalize_text)
    master_df[COL_PROJECT] = master_df[COL_PROJECT].fillna("Unmapped").astype(str).str.strip()

    # ---------------------------
    # Business Group filter first
    # ---------------------------
    if selected_business_group == "All":
        master_df = master_df[
            master_df[COL_BUSINESS_GROUP].str.upper().isin(
                [bg.upper() for bg in ALLOWED_BUSINESS_GROUPS]
            )
        ].copy()
    else:
        master_df = master_df[
            master_df[COL_BUSINESS_GROUP].str.upper()
            == selected_business_group.upper()
        ].copy()

    # ---------------------------
    # Skill Type filter
    # ---------------------------
    master_df = master_df[
        master_df[COL_SKILL_TYPE].str.upper()
        == selected_skill_type.upper()
    ].copy()

    # ---------------------------
    # Load result
    # ---------------------------
    result_df = read_result_file(result_source)

    result_df[RESULT_COL_EID] = result_df[RESULT_COL_EID].apply(normalize_text).str.lower()
    result_df[RESULT_COL_SKILL] = result_df[RESULT_COL_SKILL].apply(normalize_text)

    result_keep_cols = [
        RESULT_COL_EID,
        RESULT_COL_SKILL,
        RESULT_COL_PROFICIENCY,
    ]

    if RESULT_COL_PROFICIENCY_DESC in result_df.columns:
        result_keep_cols.append(RESULT_COL_PROFICIENCY_DESC)

    result_df = result_df[result_keep_cols].copy()

    result_df = result_df.rename(
        columns={
            RESULT_COL_EID: COL_EID,
            RESULT_COL_SKILL: COL_SKILL_NAME,
            RESULT_COL_PROFICIENCY: "Result Proficiency",
            RESULT_COL_PROFICIENCY_DESC: "Result Proficiency Description",
        }
    )

    result_df = result_df.drop_duplicates(
        subset=[COL_EID, COL_SKILL_NAME],
        keep="last",
    )

    # ---------------------------
    # Merge
    # ---------------------------
    
    merged_df = master_df.merge(
    result_df,
    on=[COL_EID, COL_SKILL_NAME],
    how="left",
)

    # ============================================================
    # ASSESSMENT + COMPETENCY LOGIC
    # ============================================================


    proficiency_map = {
        "NULL": -1,
        "P0": 0,
        "P1": 1,
        "P2": 2,
        "P3": 3,
        "Expert eligible": 4,
    }

    merged_df["proficiency_num"] = (
    merged_df["Result Proficiency Description"]
        .fillna("NULL")
        .map(proficiency_map)
    )


    # ============================================================
    # Proficiency rank based on Proficiency Description
    # Source of truth:
    # NULL = No Assessment
    # P0 = 0
    # P1 = 1
    # P2 = 2
    # P3 = 3
    # Expert eligible = 4
    # ============================================================

    PROFICIENCY_RANK_MAP = {
        "P0": 0,
        "P1": 1,
        "P2": 2,
        "P3": 3,
        "Expert eligible": 4,
    }

    merged_df["proficiency_desc_clean"] = (
        merged_df["Result Proficiency Description"]
        .astype(str)
        .str.strip()
    )

    merged_df.loc[
        merged_df["proficiency_desc_clean"].str.upper().isin(["NULL", "NAN", "NONE", ""]),
        "proficiency_desc_clean"
    ] = None

    merged_df["proficiency_num"] = (
        merged_df["proficiency_desc_clean"]
        .map(PROFICIENCY_RANK_MAP)
    )

    merged_df["has_assessment"] = merged_df["proficiency_num"].notna()


    def get_target_proficiency(management_level):
        level = pd.to_numeric(
            management_level,
            errors="coerce"
        )

        if pd.isna(level):
            return None

        level = int(level)

        # Business rule:
        # CL11/CL12 target = P2
        # CL10 and up target = P3

        if level in [11, 12]:
            return 2   # P2

        if level <= 10:
            return 3   # P3

        return None

    


    merged_df["target_proficiency_num"] = (
        merged_df[COL_LEVEL]
        .apply(get_target_proficiency)
    )

    merged_df["meets_target"] = (
        merged_df["has_assessment"]
        & merged_df["target_proficiency_num"].notna()
        & (
            merged_df["proficiency_num"]
            >= merged_df["target_proficiency_num"]
        )
    )

    merged_df["below_target"] = (
        merged_df["has_assessment"]
        & merged_df["target_proficiency_num"].notna()
        & (
            merged_df["proficiency_num"]
            < merged_df["target_proficiency_num"]
        )
    )

    merged_df["Action Reason"] = "Meeting Target"

    merged_df.loc[
        merged_df["has_assessment"] == False,
        "Action Reason"
    ] = "No Assessment"

    merged_df.loc[
        merged_df["below_target"] == True,
        "Action Reason"
    ] = "Below Target"

    # ---------------------------
    # Resource-level
    # ---------------------------

    resource_df = (
        merged_df
        .groupby(COL_PERSONNEL_NO, as_index=False)
        .agg(
            EID=(COL_EID, "first"),
            BusinessGroup=(COL_BUSINESS_GROUP, "first"),
            ManagementLevel=(COL_LEVEL, "first"),
            Project=(COL_PROJECT, "first"),
            SkillName=(COL_SKILL_NAME, "first"),
            SkillType=(COL_SKILL_TYPE, "first"),
            HasAssessment=("has_assessment", "max"),
            MeetingTarget=("meets_target", "max"),
            BelowTarget=("below_target", "max"),
            ActionReason=("Action Reason", "first"),
            ActualProficiencyRank=("proficiency_num", "max"),
            TargetProficiencyRank=("target_proficiency_num", "first"),
        )
    )

    PROFICIENCY_LABEL_MAP = {
        0: "P0",
        1: "P1",
        2: "P2",
        3: "P3",
        4: "Expert eligible",
    }

    resource_df["ActualProficiency"] = (
        resource_df["ActualProficiencyRank"]
        .map(PROFICIENCY_LABEL_MAP)
    )

    resource_df["TargetProficiency"] = (
        resource_df["TargetProficiencyRank"]
        .map(PROFICIENCY_LABEL_MAP)
    )



    return merged_df, resource_df


# ============================================================
# STREAMLIT APP
# ============================================================

st.set_page_config(
    page_title="myCompetency Simple Scorecard",
    page_icon="📌",
    layout="wide",
)

page_titles = {
    "Primary Scorecard": "📌 myCompetency Simple Scorecard",
    "Secondary Skills Explorer": "🚀 myCompetency Secondary Skills Explorer",
}
page_captions = {
    "Primary Scorecard":
        "Capability visibility dashboard showing assessment completion, compliance, intervention opportunities, and workforce readiness indicators.",

    "Secondary Skills Explorer":
        "Discover adjacent skills, hidden talent pools, and capability pathways to accelerate readiness for strategic offerings.",

    "Historical Executive Summary":
        "Historical trends, target projections, and readiness progress tracking across capability snapshots.",
}
# ---------------------------
# Sidebar
# ---------------------------

master_file = st.sidebar.file_uploader(
    "Upload Master Source File",
    type=["xlsx"],
)

result_file = st.sidebar.file_uploader(
    "Upload Proficiency Result File",
    type=["xlsx"],
)

selected_business_group = st.sidebar.selectbox(
    "Business Group",
    ["All"] + ALLOWED_BUSINESS_GROUPS,
    index=0,
)

selected_page = st.sidebar.radio(
    "Navigation",
    ["Primary Scorecard", "Secondary Skills Explorer"],
    index=0,
)

min_project_resources = st.sidebar.number_input(
    "Minimum project resources",
    min_value=1,
    value=5,
    step=1,
)

show_eid = st.sidebar.checkbox("Show EID (lowercase)", value=False)

st.title(page_titles.get(selected_page, page_titles["Primary Scorecard"]))
st.caption(page_captions.get(selected_page, page_captions["Primary Scorecard"]))

master_source = (
    master_file.getvalue()
    if master_file is not None
    else DEFAULT_MASTER_FILE
)

result_source = (
    result_file.getvalue()
    if result_file is not None
    else DEFAULT_RESULT_FILE
)

if selected_page == "Secondary Skills Explorer":
    # Provide upload widgets in the main app sidebar and pass them to the renderer
    uploaded_skills = st.sidebar.file_uploader(
        "1. Skills Dump",
        type=["xlsx"],
        key="sec_skills_upload",
    )
    uploaded_target = st.sidebar.file_uploader(
        "2. Career Level Target Lookup",
        type=["xlsx"],
        key="sec_target_upload",
    )
    uploaded_project = st.sidebar.file_uploader(
        "3. Project Lookup",
        type=["xlsx"],
        key="sec_project_upload",
    )

    render_secondary_dashboard(
        skills_source=uploaded_skills,
        target_source=uploaded_target,
        project_source=uploaded_project,
        selected_business_group=selected_business_group,
        show_business_group_select=False,
        set_page_config=False,
    )
    st.stop()

selected_skill_type = "Primary"

try:
    merged_df, resource_df = build_data(
        master_source,
        result_source,
        selected_business_group,
        selected_skill_type,
    )

except Exception as e:
    st.error("Failed to load input files.")
    st.exception(e)
    st.stop()


# ============================================================
# SCORECARD
# ============================================================

total_resources = resource_df[COL_PERSONNEL_NO].nunique()

assessed_resources = resource_df.loc[
    resource_df["HasAssessment"] == True,
    COL_PERSONNEL_NO,
].nunique()

meeting_target_resources = resource_df.loc[
    resource_df["MeetingTarget"] == True,
    COL_PERSONNEL_NO,
].nunique()

below_target_resources = resource_df.loc[
    resource_df["BelowTarget"] == True,
    COL_PERSONNEL_NO,
].nunique()

no_assessment = total_resources - assessed_resources

completion_pct = (
    assessed_resources / total_resources * 100
    if total_resources > 0
    else 0
)

target_compliance_pct = (
    meeting_target_resources / total_resources * 100
    if total_resources > 0
    else 0
)

scope_label = (
    "Tech_Song + Tech_Adobe Platform"
    if selected_business_group == "All"
    else selected_business_group
)

st.subheader(
    f"Overall {scope_label} {selected_skill_type} Skill Scorecard"
)

col1, col2, col3, col4, col5, col6 = st.columns(6)

col1.metric(
    "Total Resources",
    f"{total_resources:,}"
)

col2.metric(
    "Assessed Resources",
    f"{assessed_resources:,}"
)

col3.metric(
    "Completion %",
    f"{completion_pct:.1f}%"
)

col4.metric(
    "Target Compliance %",
    f"{target_compliance_pct:.1f}%"
)

col5.metric(
    "No Assessment",
    f"{no_assessment:,}"
)

col6.metric(
    "Below Target",
    f"{below_target_resources:,}"
)


# ============================================================
# CAREER LEVEL COMPETENCY HEALTH

st.subheader("Career Level Competency Health")

# Attempt to parse numeric career level from ManagementLevel
resource_df["career_level_num"] = pd.to_numeric(
    resource_df["ManagementLevel"],
    errors="coerce",
)

career_summary = (
    resource_df
    .groupby("career_level_num", as_index=False)
    .agg(
        TotalResources=(COL_PERSONNEL_NO, "nunique"),
        AssessedResources=("HasAssessment", "sum"),
        MeetingTarget=("MeetingTarget", "sum"),
        BelowTarget=("BelowTarget", "sum"),
    )
    .sort_values("career_level_num", ascending=False)
)

career_summary["No Assessment"] = (
    career_summary["TotalResources"] - career_summary["AssessedResources"]
)

career_summary["Completion %"] = (
    career_summary["AssessedResources"] / career_summary["TotalResources"] * 100
).fillna(0).round(0)

career_summary["Target Compliance %"] = (
    career_summary["MeetingTarget"] / career_summary["TotalResources"] * 100
).fillna(0).round(0)

career_summary = career_summary.rename(
    columns={
        "career_level_num": "Career Level",
        "TotalResources": "Total Resources",
        "BelowTarget": "Below Target",
    }
)

st.caption(
    "No Assessment indicates completion gaps. Below Target indicates assessed resources that did not meet the target proficiency."
)

st.dataframe(
    career_summary[
        [
            "Career Level",
            "Total Resources",
            "No Assessment",
            "Completion %",
            "Target Compliance %",
            "Below Target",
        ]
    ],
    use_container_width=True,
    hide_index=True,
)


# ============================================================
# DETAILS
# ============================================================

display_df = resource_df.rename(
    columns={
        "ManagementLevel": "Management Level",
        "HasAssessment": "Has Assessment",
    }
)

# Option to show/hide the filtered resource detail (hidden by default to speed UI)
show_resource_detail = st.sidebar.checkbox("Show Resource Assessment Detail", value=False)

if show_resource_detail:
    st.subheader("Filtered Resource Chase Detail")

    filtered_display = display_df.copy()

    # Normalize some column names for display / export parity with app.py
    filtered_display = filtered_display.rename(
        columns={
            "Enterpriseid": "EID",
            "SkillName": "Primary Skill",
            "career_level_num": "Career Level",
        }
    )

    cols_to_show = [
        "EID",
        "Project",
        "Primary Skill",
        "Career Level",
        "Target",
        "Actual",
        "Action Reason",
    ]

    cols_present = [c for c in cols_to_show if c in filtered_display.columns]

    # Ensure EID values are lowercase for exports/display
    if "EID" in filtered_display.columns:
        filtered_display["EID"] = filtered_display["EID"].astype(str).str.lower()

    # Default behavior: show Personnel No only. If `show_eid` is checked, show EID only.
    if show_eid and "EID" in filtered_display.columns:
        cols_eid_only = [c for c in ["EID", "Project", "Primary Skill", "Career Level", "Target", "Actual", "Action Reason"] if c in filtered_display.columns]
        st.dataframe(
            filtered_display[cols_eid_only],
            use_container_width=True,
            hide_index=True,
        )

        csv_data = filtered_display[cols_eid_only].to_csv(index=False).encode("utf-8")
        
    else:
        # Show Personnel No (COL_PERSONNEL_NO) when EID is not toggled or missing
        if COL_PERSONNEL_NO not in filtered_display.columns and COL_PERSONNEL_NO in display_df.columns:
            filtered_display[COL_PERSONNEL_NO] = display_df[COL_PERSONNEL_NO]

        cols_personnel_only = [c for c in [COL_PERSONNEL_NO, "Project", "Primary Skill", "Career Level", "Target", "Actual", "Action Reason"] if c in filtered_display.columns]
        st.dataframe(
            filtered_display[cols_personnel_only],
            use_container_width=True,
            hide_index=True,
        )

        csv_data = filtered_display[cols_personnel_only].to_csv(index=False).encode("utf-8")

    st.download_button(
        "Download filtered chase list as CSV",
        data=csv_data,
        file_name="filtered_mycompetency_chase_list.csv",
        mime="text/csv",
    )


# ============================================================
# PROJECT ACTION LIST (aggregated by EID instead of Personnel No)
# ============================================================

st.subheader("Project Action / Chase List")

@st.cache_data
def build_project_rank(resource_df, min_project_resources):
    project_summary_rows = []

    for project_name, group in resource_df.groupby("Project"):
        project_total_resources = group["EID"].nunique()
        project_assessed_resources = group.loc[group["HasAssessment"] == True, "EID"].nunique()
        project_no_assessment = group.loc[group["HasAssessment"] == False, "EID"].nunique()
        project_below_target_only = group.loc[
            (group["BelowTarget"] == True) & (group["HasAssessment"] == True),
            "EID",
        ].nunique()
        project_meeting_target = group.loc[group["MeetingTarget"] == True, "EID"].nunique()
        project_resources_to_chase = project_no_assessment + project_below_target_only

        project_summary_rows.append(
            {
                "Project": project_name,
                "TotalResources": project_total_resources,
                "AssessedResources": project_assessed_resources,
                "No Assessment": project_no_assessment,
                "Below Target Only": project_below_target_only,
                "Meeting Target": project_meeting_target,
                "Resources To Chase": project_resources_to_chase,
                "Chase %": (project_resources_to_chase / project_total_resources * 100) if project_total_resources > 0 else 0,
                "Completion %": (project_assessed_resources / project_total_resources * 100) if project_total_resources > 0 else 0,
                "Target Compliance %": (project_meeting_target / project_total_resources * 100) if project_total_resources > 0 else 0,
                "Priority Score": project_no_assessment + (project_below_target_only * 2),
            }
        )

    project_view = pd.DataFrame(project_summary_rows)
    if len(project_view) > 0:
        project_view = project_view.sort_values(["Resources To Chase", "TotalResources"], ascending=[False, False])

    return project_view[project_view["TotalResources"] >= min_project_resources].copy()

eligible_projects = build_project_rank(resource_df, min_project_resources)

@st.cache_data
def build_drilldown_df(resource_df, selected_project):
    drilldown_df = resource_df[resource_df["Project"] == selected_project].copy()
    # Create a stable Action Reason ordering and sort by Action Reason then Career Level (high->low)
    # Work on a copy and ensure the Career Level numeric column is available for sorting
    if "ActionReason" in drilldown_df.columns:
        action_order = ["No Assessment", "Below Target", "Meeting Target"]
        drilldown_df["Action Reason Order"] = pd.Categorical(
            drilldown_df["ActionReason"],
            categories=action_order,
            ordered=True,
        )

    # ensure we have a numeric career level to sort by (higher first)
    if "career_level_num" in drilldown_df.columns:
        drilldown_df["_career_level_sort"] = pd.to_numeric(drilldown_df["career_level_num"], errors="coerce")
    elif "Career Level" in drilldown_df.columns:
        drilldown_df["_career_level_sort"] = pd.to_numeric(drilldown_df["Career Level"], errors="coerce")
    else:
        drilldown_df["_career_level_sort"] = pd.NA

    # Build sort keys: Action Reason (asc per mapping), Career Level (desc), then Primary Skill, then Personnel No for stability
    sort_by = []
    sort_asc = []
    if "Action Reason Order" in drilldown_df.columns:
        sort_by.append("Action Reason Order")
        sort_asc.append(True)

    sort_by.append("_career_level_sort")
    sort_asc.append(True)

    if "SkillName" in drilldown_df.columns:
        sort_by.append("SkillName")
        sort_asc.append(True)
    elif "Primary Skill" in drilldown_df.columns:
        sort_by.append("Primary Skill")
        sort_asc.append(True)

    if COL_PERSONNEL_NO in drilldown_df.columns:
        sort_by.append(COL_PERSONNEL_NO)
        sort_asc.append(True)

    # Perform sort; place NaNs last
    drilldown_df = drilldown_df.sort_values(by=sort_by, ascending=sort_asc, na_position="last")

    # Rename for display parity
    drilldown_df = drilldown_df.rename(
        columns={
            "SkillName": "Primary Skill",
            "ManagementLevel": "Management Level",
            "ActualProficiency": "Actual Proficiency",
            "TargetProficiency": "Target Proficiency",
            "career_level_num": "Career Level",
            "ActionReason": "Action Reason",
            "EID": "EID",
        }
    )

    drilldown_cols = [
        COL_PERSONNEL_NO,
        "EID",
        "Project",
        "Primary Skill",
        "Career Level",
        "Target Proficiency",
        "Actual Proficiency",
        "Action Reason",
    ]

    drilldown_cols = [c for c in drilldown_cols if c in drilldown_df.columns]

    # clean up helper sort column
    if "Action Reason Order" in drilldown_df.columns:
        drilldown_df = drilldown_df.drop(columns=["Action Reason Order"])
    if "_career_level_sort" in drilldown_df.columns:
        drilldown_df = drilldown_df.drop(columns=["_career_level_sort"])

    return drilldown_df[drilldown_cols]

sort_option = st.radio(
    "Sort projects by",
    [
        "Resources To Chase",
        "Chase %",
        "Priority Score",
        "Lowest Completion %",
        "Lowest Target Compliance %",
    ],
    horizontal=True,
)

if len(eligible_projects) == 0:
    st.warning("No projects met the current minimum resource threshold.")
    project_rank = eligible_projects.copy()
else:
    if sort_option == "Resources To Chase":
        project_rank = eligible_projects.sort_values(["Resources To Chase", "TotalResources"], ascending=[False, False])
    elif sort_option == "Chase %":
        project_rank = eligible_projects.sort_values(["Chase %", "TotalResources"], ascending=[False, False])
    elif sort_option == "Priority Score":
        project_rank = eligible_projects.sort_values(["Priority Score", "TotalResources"], ascending=[False, False])
    elif sort_option == "Lowest Completion %":
        project_rank = eligible_projects.sort_values(["Completion %", "TotalResources"], ascending=[True, False])
    else:
        project_rank = eligible_projects.sort_values(["Target Compliance %", "TotalResources"], ascending=[True, False])

    display_cols = [
        "Project",
        "TotalResources",
        "No Assessment",
        "Below Target Only",
        "Resources To Chase",
        "Chase %",
        "Completion %",
        "Target Compliance %",
        "Priority Score",
    ]

    project_display_df = project_rank.reindex(columns=display_cols).round(0).head(50)

    st.dataframe(
        project_display_df,
        width="stretch",
        use_container_width=True,
        hide_index=True,
    )


# ============================================================
# PROJECT DRILLDOWN: Who is in this project? (EID-based)
# ============================================================

st.subheader("Project Drilldown: Who is in this project?")

if len(project_rank) > 0:
    project_dropdown_options = project_rank["Project"].dropna().unique().tolist()

    selected_drilldown_project = st.selectbox(
        "Select project for drilldown",
        project_dropdown_options,
    )

    drilldown_df = build_drilldown_df(resource_df, selected_drilldown_project)

    # Ensure EID display is lowercase and available
    if "EID" in drilldown_df.columns:
        drilldown_df["EID"] = drilldown_df["EID"].astype(str).str.lower()

    # Choose which identifier column to show based on sidebar toggle
    if show_eid and "EID" in drilldown_df.columns:
        display_cols = [c for c in ["EID", "Project", "Primary Skill", "Career Level", "Target Proficiency", "Actual Proficiency", "Action Reason"] if c in drilldown_df.columns]
    else:
        # Prefer Personnel No when EID is hidden
        if COL_PERSONNEL_NO in drilldown_df.columns:
            # ensure Personnel No column exists as-is
            pass

        display_cols = [c for c in [COL_PERSONNEL_NO, "Project", "Primary Skill", "Career Level", "Target Proficiency", "Actual Proficiency", "Action Reason"] if c in drilldown_df.columns]

    st.dataframe(
        drilldown_df[display_cols],
        use_container_width=True,
        hide_index=True,
    )

else:
    st.warning("No project available for drilldown.")


# ============================================================
# SKILL GAP ANALYSIS
# ============================================================

st.subheader("Primary Skill Gap Analysis")

skill_source = resource_df.copy()

# Aggregation by SkillName
skill_gap = (
    skill_source
    .groupby("SkillName", as_index=False)
    .agg(
        TotalResources=(COL_PERSONNEL_NO, "nunique"),
        NoAssessment=(
            COL_PERSONNEL_NO,
            lambda s: skill_source.loc[s.index][skill_source.loc[s.index, "HasAssessment"] == False][COL_PERSONNEL_NO].nunique(),
        ),
        BelowTargetOnly=(
            COL_PERSONNEL_NO,
            lambda s: skill_source.loc[s.index][(skill_source.loc[s.index, "BelowTarget"] == True) & (skill_source.loc[s.index, "HasAssessment"] == True)][COL_PERSONNEL_NO].nunique(),
        ),
        MeetingTarget=(
            COL_PERSONNEL_NO,
            lambda s: skill_source.loc[s.index][skill_source.loc[s.index, "MeetingTarget"] == True][COL_PERSONNEL_NO].nunique(),
        ),
    )
)

skill_gap["Resources To Chase"] = (
    skill_gap["NoAssessment"] + skill_gap["BelowTargetOnly"]
)

skill_gap["Target Gap %"] = (
    skill_gap["Resources To Chase"] / skill_gap["TotalResources"] * 100
)

skill_gap["Priority Score"] = (
    skill_gap["NoAssessment"] + (skill_gap["BelowTargetOnly"] * 2)
)

skill_sort = st.radio(
    "Sort skills by",
    [
        "Resources To Chase",
        "NoAssessment",
        "BelowTargetOnly",
        "Priority Score",
        "Target Gap %",
    ],
    horizontal=True,
)

skill_gap_rank = (
    skill_gap
    .sort_values([skill_sort, "TotalResources"], ascending=[False, False])
    .head(30)
)

c1, c2 = st.columns(2)

with c1:
    st.dataframe(
        skill_gap_rank[
            [
                "SkillName",
                "TotalResources",
                "NoAssessment",
                "BelowTargetOnly",
                "Resources To Chase",
                "Target Gap %",
                "Priority Score",
            ]
        ],
        use_container_width=True,
        hide_index=True,
    )

with c2:
    fig_skill = px.bar(
        skill_gap_rank.sort_values("Resources To Chase", ascending=True),
        x="Resources To Chase",
        y="SkillName",
        orientation="h",
        title="Top Primary Skills With Most Resources To Chase",
        hover_data=["NoAssessment", "BelowTargetOnly", "TotalResources"],
    )

    st.plotly_chart(fig_skill, use_container_width=True)


# ============================================================
# PER SKILL SUMMARY
# ============================================================

skill_summary_df = (
    merged_df
    .groupby([COL_SKILL_NAME, COL_SKILL_TYPE], as_index=False)
    .agg(
        Total_Resources=(COL_PERSONNEL_NO, "nunique"),
        Assessed_Resources=(
            COL_PERSONNEL_NO,
            lambda x: merged_df.loc[
                x.index,
                "has_assessment"
            ].eq(True).groupby(x).any().sum()
        ),
        Meeting_Target=(
            COL_PERSONNEL_NO,
            lambda x: merged_df.loc[
                x.index,
                "meets_target"
            ].eq(True).groupby(x).any().sum()
        ),
        Below_Target=(
            COL_PERSONNEL_NO,
            lambda x: merged_df.loc[
                x.index,
                "below_target"
            ].eq(True).groupby(x).any().sum()
        ),
    )
)

skill_summary_df["No_Assessment"] = (
    skill_summary_df["Total_Resources"]
    - skill_summary_df["Assessed_Resources"]
)

skill_summary_df["Completion %"] = (
    skill_summary_df["Assessed_Resources"]
    / skill_summary_df["Total_Resources"]
    * 100
).round(1)

skill_summary_df["Compliance %"] = (
    skill_summary_df["Meeting_Target"]
    / skill_summary_df["Total_Resources"]
    * 100
).round(1)

skill_summary_df = skill_summary_df.sort_values(
    by="Total_Resources",
    ascending=False
)

st.subheader("Per Skill Summary")

st.dataframe(
    skill_summary_df,
    use_container_width=True
)





# ============================================================
# SKILL DRILLDOWN: WHO IS IN THIS SKILL?
# ============================================================

st.subheader("Skill Drilldown: Who is in this skill?")

skill_options = (
    skill_summary_df
    .apply(
        lambda x:
        f"{x[COL_SKILL_NAME]} "
        f"(Resources: {x['Total_Resources']}, "
        f"No Assessment: {x['No_Assessment']}, "
        f"Below Target: {x['Below_Target']})",
        axis=1
    )
    .tolist()
)

selected_skill_option = st.selectbox(
    "Select skill for drilldown",
    skill_options,
    key="skill_drilldown"
)

selected_skill = selected_skill_option.split(" (")[0]

skill_members_df = (
    merged_df[
        merged_df[COL_SKILL_NAME] == selected_skill
    ]
    .copy()
)

skill_members_df = skill_members_df.merge(
    resource_df[
        [
            COL_PERSONNEL_NO,
            "EID",
            "Project",
            "ManagementLevel",
            "ActionReason",
            "ActualProficiency",
            "TargetProficiency",
        ]
    ],
    on=COL_PERSONNEL_NO,
    how="left"
)

# Keep one row per resource
skill_members_df = (
    skill_members_df
    .drop_duplicates(
        subset=[COL_PERSONNEL_NO]
    )
    .copy()
)

# Keep only resources to chase
skill_members_df = skill_members_df[
    skill_members_df["ActionReason"].isin(
        [
            "No Assessment",
            "Below Target",
        ]
    )
].copy()

# Sort Action Reason first
action_order = {
    "No Assessment": 1,
    "Below Target": 2,
}

skill_members_df["ActionOrder"] = (
    skill_members_df["ActionReason"]
    .map(action_order)
    .fillna(999)
)

# Sort Level properly
skill_members_df["LevelSort"] = pd.to_numeric(
    skill_members_df["ManagementLevel"],
    errors="coerce"
)

skill_members_df = skill_members_df.sort_values(
    by=[
        "ActionOrder",
        "LevelSort",
        COL_PERSONNEL_NO,
    ]
)

# Decide identifier column
if show_eid and "EID" in skill_members_df.columns:
    identifier_col = "EID"
else:
    identifier_col = COL_PERSONNEL_NO

display_cols = [
    c
    for c in [
        identifier_col,
        "Project",
        "ManagementLevel",
        "TargetProficiency",
        "ActualProficiency",
        "ActionReason",
    ]
    if c in skill_members_df.columns
]

display_df = skill_members_df[display_cols].copy()

display_df = display_df.rename(
    columns={
         "Project": "Project",
        "ManagementLevel": "Level",
        "TargetProficiency": "Target Proficiency",
        "ActualProficiency": "Actual Proficiency",
        "ActionReason": "Action Reason",
    }
)

st.dataframe(
    display_df,
    use_container_width=True,
    hide_index=True,
)

# ============================================================
# CSV DOWNLOAD
# ============================================================

safe_skill = (
    selected_skill
    .replace("/", "_")
    .replace("\\", "_")
    .replace(" ", "_")
)

csv_data = display_df.to_csv(
    index=False
).encode("utf-8")

st.download_button(
    label="Download Skill Drilldown CSV",
    data=csv_data,
    file_name=f"myCompetency_{safe_skill}_Skill_Drilldown.csv",
    mime="text/csv"
)


import io
import re

def safe_sheet_name(name):
    # Excel sheet names max 31 chars and cannot contain these: \ / ? * [ ]
    cleaned = re.sub(r'[\\/*?:\[\]]', "_", str(name))
    return cleaned[:31]


# ============================================================
# DOWNLOAD: ONE EXCEL FILE, ONE SHEET PER SKILL
# ============================================================

def build_skill_chase_workbook(
    skill_summary_df,
    merged_df,
    resource_df,
    show_eid,
):
    output = io.BytesIO()

    action_order = {
        "No Assessment": 1,
        "Below Target": 2,
        "Meeting Target": 3,
    }

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ====================================================
        # Sheet 1: Per Skill Summary
        # ====================================================

        skill_summary_df.to_excel(
            writer,
            sheet_name="Per Skill Summary",
            index=False
        )

        # ====================================================
        # One sheet per skill
        # ====================================================

        for _, skill_row in skill_summary_df.iterrows():

            skill_name = skill_row[COL_SKILL_NAME]

            skill_members_df = (
                merged_df[
                    merged_df[COL_SKILL_NAME] == skill_name
                ]
                .copy()
            )

            # Merge resource-level fields from resource_df
            skill_members_df = skill_members_df.merge(
                resource_df[
                    [
                        COL_PERSONNEL_NO,
                        "EID",
                        "Project",
                        "ManagementLevel",
                        "ActionReason",
                        "ActualProficiency",
                        "TargetProficiency",
                    ]
                ],
                on=COL_PERSONNEL_NO,
                how="left"
            )

            # One row per resource
            skill_members_df = (
                skill_members_df
                .drop_duplicates(
                    subset=[COL_PERSONNEL_NO]
                )
                .copy()
            )

            # Keep only resources to chase
            skill_members_df = skill_members_df[
                skill_members_df["ActionReason"].isin(
                    [
                        "No Assessment",
                        "Below Target",
                    ]
                )
            ].copy()

            # Skip skills with no resources to chase
            if skill_members_df.empty:
                continue

            # Sorting
            skill_members_df["ActionOrder"] = (
                skill_members_df["ActionReason"]
                .map(action_order)
                .fillna(999)
            )

            skill_members_df["LevelSort"] = pd.to_numeric(
                skill_members_df["ManagementLevel"],
                errors="coerce"
            )

            skill_members_df = skill_members_df.sort_values(
                by=[
                    "ActionOrder",
                    "LevelSort",
                    COL_PERSONNEL_NO,
                ],
                ascending=[
                    True,
                    True,
                    True,
                ],
                na_position="last"
            )

            # Decide identifier column
            if show_eid and "EID" in skill_members_df.columns:
                identifier_col = "EID"
            else:
                identifier_col = COL_PERSONNEL_NO

            display_cols = [
                c
                for c in [
                    identifier_col,
                    "Project",
                    "ManagementLevel",
                    "TargetProficiency",
                    "ActualProficiency",
                    "ActionReason",
                ]
                if c in skill_members_df.columns
            ]

            export_df = skill_members_df[display_cols].copy()

            export_df = export_df.rename(
                columns={
                    "EID": "EID",
                    COL_PERSONNEL_NO: "Personnel No",
                    "ManagementLevel": "Level",
                    "TargetProficiency": "Target Proficiency",
                    "ActualProficiency": "Actual Proficiency",
                    "ActionReason": "Action Reason",
                }
            )

            sheet_name = safe_sheet_name(skill_name)

            export_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

        # ====================================================
        # Basic formatting
        # ====================================================

        workbook = writer.book

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            ws.freeze_panes = "A2"

            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    try:
                        cell_length = len(str(cell.value)) if cell.value is not None else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except Exception:
                        pass

                ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    output.seek(0)
    return output

cols = [
    COL_PERSONNEL_NO,
    "EID",
    "Project",
    "ManagementLevel",
    "ActionReason",
    "ActualProficiency",
    "TargetProficiency",
]


skill_chase_excel = build_skill_chase_workbook(
    skill_summary_df=skill_summary_df,
    merged_df=merged_df,
    resource_df=resource_df,
    show_eid=show_eid,
)

st.download_button(
    label="Download Skill Chase Workbook",
    data=skill_chase_excel,
    file_name="myCompetency_Skill_Chase_Workbook.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# ============================================================
# EXCEL OUTPUT FOR PEOPLE LEAD / PROJECT FOLLOW-UP
# ============================================================

def create_chase_excel(
    summary_df,
    project_df,
    resource_detail_df,
    skill_gap_df,
    career_summary_df,
    selected_business_group,
    assessment_scope,
):
    output = io.BytesIO()

    # local imports used by the full writer
    import matplotlib.pyplot as plt
    from openpyxl.drawing.image import Image as ExcelImage

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write sheets
        summary_df.to_excel(
            writer,
            sheet_name="Executive Summary",
            index=False,
            startrow=0,
        )

        project_df.to_excel(
            writer,
            sheet_name="Project Action Summary",
            index=False,
        )

        resource_detail_df.to_excel(
            writer,
            sheet_name="Resource Chase Detail",
            index=False,
        )

        skill_gap_df.to_excel(
            writer,
            sheet_name="Skill Gap Summary",
            index=False,
        )

        career_summary_df.to_excel(
            writer,
            sheet_name="Career Level Health",
            index=False,
        )

        workbook = writer.book
        ws = writer.sheets["Executive Summary"]
        from openpyxl.styles import Font, PatternFill, Alignment

        # Style Executive Summary row 1 and row 2
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

        value_fill = PatternFill(fill_type="solid", fgColor="F3F8FC")

        for cell in ws[1]:
            cell.font = Font(bold=True, size=13)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[2]:
            cell.font = Font(bold=True, size=13)
            cell.fill = value_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 24

        # Add metric definitions in same Executive Summary sheet
        definitions = [
            ["Metric", "Meaning"],
            ["Business Group", "Business group included in the analysis, such as Tech_Song, Tech_Adobe Platform, or All."],
            ["Assessment Scope", "Indicates whether the analysis includes Primary Skills, Secondary Skills, or All Skills."],
            ["Total Resources", "Total unique resources included in the selected business group and assessment scope."],
            ["Assessed Resources", "Resources with a completed competency assessment."],
            ["Completion %", "Percentage of resources with completed assessments. Formula: Assessed Resources / Total Resources x 100."],
            ["Target Compliance %", "Percentage of resources meeting or exceeding the required target proficiency."],
            ["No Assessment", "Resources without a completed competency assessment. These require assessment completion follow-up."],
            ["Below Target", "Resources who completed an assessment but are below the required target proficiency. These require capability uplift, learning, coaching, or reassessment action."],
        ]

        # Put definitions starting row 5 para hindi matatamaan yung summary
        start_row = 5

        for r_idx, row in enumerate(definitions, start=start_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Create Top Projects chart as image
        chart_start_row = start_row + len(definitions) + 3

        try:
            top_projects = (
                project_df
                .sort_values("Resources To Chase", ascending=False)
                .head(10)
                .copy()
            )

            if len(top_projects) > 0:
                plt.figure(figsize=(10, 6))
                plt.barh(
                    top_projects["Project"],
                    top_projects["Resources To Chase"],
                )
                plt.xlabel("Resources To Chase")
                plt.ylabel("Project")
                plt.title("Top Projects With Most Resources To Chase")
                plt.gca().invert_yaxis()
                plt.tight_layout()

                chart_path = "top_projects_chart.png"
                plt.savefig(chart_path, dpi=150)
                plt.close()

                img = ExcelImage(chart_path)
                img.width = 720
                img.height = 420

                ws.add_image(img, f"A{chart_start_row}")

        except Exception:
            # Do not fail Excel generation if chart creation fails
            ws.cell(
                row=chart_start_row,
                column=1,
                value="Chart could not be generated. Please refer to Project Action Summary tab.",
            )

        # Basic formatting
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            if sheet_name == "Resource Chase Detail":
                project_header_fill = PatternFill(
                    fill_type="solid",
                    fgColor="BDD7EE"
                )

                for row in range(2, worksheet.max_row + 1):
                    project_value = worksheet.cell(row=row, column=2).value

                    if isinstance(project_value, str) and project_value.startswith("PROJECT:"):
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.font = Font(bold=True, size=12)
                            cell.fill = project_header_fill
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                        worksheet.row_dimensions[row].height = 22
            worksheet.freeze_panes = "A2"

            # Round numeric values to no decimals and set integer number format
            for col_idx, header_cell in enumerate(worksheet[1], start=1):
                for row_idx in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        try:
                            cell.value = round(cell.value, 0)
                        except Exception:
                            pass
                        cell.number_format = '0'

            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    try:
                        cell_length = len(str(cell.value)) if cell.value is not None else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except Exception:
                        pass

                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)

    output.seek(0)
    return output


def build_grouped_resource_export(resource_export):
    grouped_rows = []
    columns = list(resource_export.columns)

    for project_name, group in resource_export.groupby("Project", sort=True):
        project_header = {col: "" for col in columns}
        project_header["Project"] = f"PROJECT: {project_name}"
        grouped_rows.append(project_header)

        # Sort rows within each project by Action Reason priority
        sort_order = ["No Assessment", "Below Target", "Meeting Target"]
        g = group.copy()
        if "Action Reason" in g.columns:
            g["Action Reason Order"] = pd.Categorical(
                g["Action Reason"], categories=sort_order, ordered=True
            )
            # Rows with specified categories appear first in the defined order; others follow
            g = g.sort_values(by=["Action Reason Order"], na_position="last")
            g = g.drop(columns=["Action Reason Order"])

        # append sorted rows
        grouped_rows.extend(g.to_dict("records"))

    return pd.DataFrame(grouped_rows, columns=columns)


# Prepare export dataframes
summary_df = pd.DataFrame(
    [
        {
            "Business Group": scope_label,
            "Assessment Scope": selected_skill_type,
            "Total Resources": total_resources,
            "Assessed Resources": assessed_resources,
            "Completion %": round(completion_pct, 0),
            "Target Compliance %": round(target_compliance_pct, 0),
            "No Assessment": no_assessment,
            "Below Target": below_target_resources,
        }
    ]
)

project_export_cols = [
    "Project",
    "TotalResources",
    "No Assessment",
    "Below Target Only",
    "Resources To Chase",
    "Chase %",
    "Completion %",
    "Target Compliance %",
    "Priority Score",
]

project_export = pd.DataFrame()
if 'project_rank' in globals():
    project_export = project_rank.reindex(columns=project_export_cols).copy()

# Resource export: use displayed names and a conservative column set
resource_export = resource_df.rename(
    columns={
        "SkillName": "Primary Skill",
        "career_level_num": "Career Level",
        "ActionReason": "Action Reason",
    }
)

# select conservative columns if present
desired_cols = ["EID", "Project", "Primary Skill", "Career Level", "Target", "Actual", "Action Reason"]
resource_export = resource_export[[c for c in desired_cols if c in resource_export.columns]]

# Enforce Action Reason ordering for exports and sort per-project by Action Reason then Career Level (high->low)
action_sort_order = {
    "No Assessment": 1,
    "Below Target": 2,
    "Meeting Target": 3,
}

if "Action Reason" in resource_export.columns:
    resource_export["Action Sort"] = (
        resource_export["Action Reason"].map(action_sort_order).fillna(99)
    )

    # ensure numeric career level sort key
    if "Career Level" in resource_export.columns:
        resource_export["_career_level_sort"] = pd.to_numeric(resource_export["Career Level"], errors="coerce")
    else:
        resource_export["_career_level_sort"] = pd.NA

    resource_export = resource_export.sort_values(
        by=["Project", "Action Sort", "_career_level_sort", "Primary Skill", "EID"],
        ascending=[True, True, True, True, True],
        na_position="last",
    )

    resource_export = resource_export.drop(columns=["Action Sort", "_career_level_sort"], errors="ignore")

skill_export = skill_gap[
    [
        "SkillName",
        "TotalResources",
        "NoAssessment",
        "BelowTargetOnly",
        "Resources To Chase",
        "Target Gap %",
        "Priority Score",
    ]
].copy()

skill_export = skill_export.rename(columns={"SkillName": "Primary Skill", "NoAssessment": "No Assessment", "BelowTargetOnly": "Below Target Only"})

career_export = career_summary[
    [
        "Career Level",
        "Total Resources",
        "No Assessment",
        "Completion %",
        "Target Compliance %",
        "Below Target",
    ]
].copy()

resource_export_grouped = build_grouped_resource_export(resource_export)

excel_output = create_chase_excel(
    summary_df=summary_df,
    project_df=project_export,
    resource_detail_df=resource_export_grouped,
    skill_gap_df=skill_export,
    career_summary_df=career_export,
    selected_business_group=scope_label,
    assessment_scope=selected_skill_type,
)

st.download_button(
    "Download People Lead Follow-up Excel",
    data=excel_output,
    file_name="mycompetency_people_lead_followup_pack.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
